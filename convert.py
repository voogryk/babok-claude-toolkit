#!/usr/bin/env python3
"""
Document converter — PDF, DOCX, DOC, XLSX → Markdown.

Usage:
    python convert.py <input-file> [output-file]

If output-file is not specified, writes to stdout.
Supports: .pdf, .docx, .doc, .xlsx, .xls
"""

import subprocess
import sys
import tempfile
from pathlib import Path


def convert_pdf(path: Path) -> str:
    """Convert PDF to markdown using pymupdf4llm."""
    import pymupdf4llm

    return pymupdf4llm.to_markdown(str(path))


def convert_docx(path: Path) -> str:
    """Convert DOCX to markdown using mammoth."""
    import mammoth

    with open(path, "rb") as f:
        result = mammoth.convert_to_markdown(f)

    md = result.value

    # Log warnings
    if result.messages:
        warnings = "\n".join(f"- {m.message}" for m in result.messages)
        md += f"\n\n---\n_Conversion warnings:_\n{warnings}\n"

    return md


def convert_doc(path: Path) -> str:
    """Convert DOC to markdown via LibreOffice → DOCX → mammoth."""
    # Try antiword first (faster)
    try:
        result = subprocess.run(
            ["antiword", str(path)],
            capture_output=True,
            text=True,
            timeout=30,
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout
    except FileNotFoundError:
        pass

    # Fallback: LibreOffice convert to DOCX, then use mammoth
    with tempfile.TemporaryDirectory() as tmpdir:
        subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--convert-to",
                "docx",
                "--outdir",
                tmpdir,
                str(path),
            ],
            capture_output=True,
            timeout=60,
        )
        docx_path = Path(tmpdir) / f"{path.stem}.docx"
        if docx_path.exists():
            return convert_docx(docx_path)

    raise RuntimeError(
        f"Cannot convert {path}. Install antiword or libreoffice."
    )


def convert_xlsx(path: Path) -> str:
    """Convert XLSX to markdown tables using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        parts.append(f"## {sheet_name}\n")

        # Find max columns with data
        max_cols = max(
            len([c for c in row if c is not None]) for row in rows
        )
        if max_cols == 0:
            continue

        # Build markdown table
        for i, row in enumerate(rows):
            cells = [str(c) if c is not None else "" for c in row[:max_cols]]
            line = "| " + " | ".join(cells) + " |"
            parts.append(line)

            # Header separator after first row
            if i == 0:
                sep = "| " + " | ".join(["---"] * max_cols) + " |"
                parts.append(sep)

        parts.append("")  # blank line between sheets

    wb.close()
    return "\n".join(parts)


CONVERTERS = {
    ".pdf": convert_pdf,
    ".docx": convert_docx,
    ".doc": convert_doc,
    ".xlsx": convert_xlsx,
    ".xls": convert_xlsx,
}

SUPPORTED_EXTENSIONS = set(CONVERTERS.keys())


def convert(input_path: str) -> str:
    """Convert a document to markdown. Returns markdown string."""
    path = Path(input_path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    ext = path.suffix.lower()
    converter = CONVERTERS.get(ext)

    if not converter:
        raise ValueError(
            f"Unsupported format: {ext}. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    return converter(path)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    md = convert(input_path)

    if output_path:
        Path(output_path).write_text(md, encoding="utf-8")
        print(f"Converted: {input_path} → {output_path}", file=sys.stderr)
    else:
        print(md)


if __name__ == "__main__":
    main()
